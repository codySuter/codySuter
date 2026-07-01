#!/usr/bin/env python3
"""Build the sign-shop product dataset from STIHL dealer source files.

Inputs (place in sign-shop/tools/source/, not committed to git):
  - Dealer_Price_File.csv          STIHL dealer price export
  - STIHL_SKU_Master_Listing.xlsx  Ace/STIHL SKU master workbook

Output:
  - sign-shop/data/products.js     window.SIGN_DATA = {...}

Only customer-facing fields are emitted (MSRP, UPC, SKUs, part numbers,
descriptions). Dealer cost is intentionally excluded from the output.
"""
import csv
import json
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
SOURCE = HERE / "source"
OUT = HERE.parent / "data" / "products.js"

# Categories that are sellable power tool units (get their own signs).
UNIT_CATEGORIES = {
    "0CS": ("Gas Chain Saws", "GAS CHAIN SAW"),
    "0LB": ("Battery Chain Saws", "BATTERY CHAIN SAW"),
    "0KM": ("Kombi Powerheads", "KOMBI SYSTEM"),
    "0TR": ("Trimmers & Brushcutters", "TRIMMER / BRUSHCUTTER"),
    "0TS": ("Cut-Off Machines", "CUT-OFF MACHINE"),
    "1BB": ("Backpack Blowers", "BACKPACK BLOWER"),
    "1BH": ("Handheld Blowers", "HANDHELD BLOWER"),
    "1HB": ("Battery Hedge Trimmers", "BATTERY HEDGE TRIMMER"),
    "1HS": ("Hedge Trimmers", "HEDGE TRIMMER"),
    "1HT": ("Pole Pruners", "POLE PRUNER"),
    "1IN": ("Industrial / Augers", "EARTH AUGER"),
    "1IB": ("Battery Sprayers", "BATTERY SPRAYER"),
    "1LB": ("Battery Units", "BATTERY POWER TOOL"),
    "1MM": ("Multi-Machines", "YARD BOSS MULTI-SYSTEM"),
    "1RB": ("Pressure Washers", "PRESSURE WASHER"),
    "1RM": ("Gas Lawn Mowers", "GAS LAWN MOWER"),
    "1RZ": ("Front Mowers", "FRONT MOWER"),
    "1SE": ("Wet/Dry Vacuums", "WET/DRY VACUUM"),
    "1ZB": ("Battery Front Mowers", "BATTERY FRONT MOWER"),
}

# Words that mark the start of the product-type portion of a description.
TYPE_WORDS = (
    "Chainsaw|Cordless chain saw|Chain saw|Brushcutter|Edger|CombiEngine|"
    "Cordless KombiMotor|Pole pruner|Hedge trimmer|Hedgetrimmer|"
    "Cordless hedge trimmer|Cordless Hedgetrimmer|Blower|Cordless Blower|"
    "Mistblower|Shredder/Vacuum|Cut-off machine|Cordless cut-off machine|"
    "Earth auger|Hand held drill|Cordless trimmer|Cordless sweeper|"
    "Cordless sprayer|Electric Trimmer|Electric Blower|Electric hedge trimmer|"
    "Robotic mower|Cordless lawn mower|Lawn mower|High-pressure washer|"
    "High-pressure cleaner|Vacuums|yard boss MultiEngine|Magnum Blower"
)
TYPE_RE = re.compile(r"\s*(%s)" % TYPE_WORDS)
BAR_RE = re.compile(r"(\d+)\s*(?:cm|mm)/(\d+)\s*in")
CHAIN_RE = re.compile(r"\b(\d{2}\s?(?:RS|RM|RH|PM|PMM|PS|PD)[A-Z0-9]{0,3})\b")
NICKNAMES = ["Farm Boss", "Wood Boss", "Magnum", "yard boss", "Yard Boss"]


def dash_part(material: str) -> str:
    """'1141 200 0681 US' -> '1141-200-0681'"""
    m = material.replace(" US", "").strip()
    return m.replace(" ", "-")


BRAND_RE = re.compile(r"^i?[A-Z]{2,4}$")           # MS, MSA, iMOW, RZ, WP…
NUM_RE = re.compile(r"^\d+(\.\d+)?[a-z]?(-[A-Z])?$")  # 271, 60.0, 752.0i, 280.0-B
SUFFIX_RE = re.compile(r"^([A-Z]{1,3}(-[A-Z]{1,3})?|SET|PLUS|CONTROL|EVO)$")
SKIP_TOKENS = {"(USA)", "1/4", "3/8", "in.P"}


def extract_model(head: str) -> str:
    """Token-scan the leading product name out of a description head.

    'MSA 60.0 C-B 1/4 in.P SET Cordless chain saw' -> 'MSA 60.0 C-B SET'
    'RE 100.0 PLUS CONTROL High-pressure'          -> 'RE 100.0 PLUS CONTROL'
    """
    tokens = head.split()
    if len(tokens) < 2 or not BRAND_RE.match(tokens[0]) or not NUM_RE.match(tokens[1]):
        return head  # outliers ('Deflector', 'Mulching kit AMK 056.0') pass through
    kept = tokens[:2]
    for i in range(2, len(tokens)):
        t = tokens[i]
        if t in SKIP_TOKENS:
            continue
        # bare integer followed by an inch marker is a size spec, not the name
        if t.isdigit() and i + 1 < len(tokens) and tokens[i + 1].startswith("in"):
            break
        if SUFFIX_RE.match(t) or NUM_RE.match(t):
            kept.append(t)
            continue
        break
    return " ".join(kept)


def parse_unit(desc: str):
    """Split 'MS 271-Z Chainsaw,45cm/18 in.,23RM3' into parts."""
    m = TYPE_RE.search(desc)
    if m:
        head, ptype, tail = desc[: m.start()], m.group(1), desc[m.end():]
    else:
        head, ptype, tail = desc, "", ""
    head = head.strip().rstrip(",")
    # Drop dealer config suffixes: trailing 'Z', '-Z', '-AZ', 'LZ' tokens.
    head = re.sub(r"[- ](?:A?Z|LZ)$", "", head).strip()
    model = extract_model(head)
    bar_in = None
    bm = BAR_RE.search(desc)
    if bm:
        bar_in = int(bm.group(2))
    chain = None
    cm = CHAIN_RE.search(tail)
    if cm:
        raw = cm.group(1).replace(" ", "")
        # '23RM3' -> '23 RM3', '61PMM3' -> '61 PMM3'
        chain = re.sub(r"^(\d{2})", r"\1 ", raw)
    return model, ptype, bar_in, chain


def load_master(path: Path):
    """UPC -> {aceSku, retail, status} from the Master SKU List sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb["Master SKU List"]
    upc_map = {}
    header = None
    for row in ws.iter_rows(values_only=True):
        vals = ["" if v is None else str(v).strip() for v in row]
        if header is None:
            if vals[0] == "UPC Code":
                header = vals
            continue
        rec = dict(zip(header, vals))
        upc = rec.get("UPC Code", "")
        if upc:
            upc_map[upc] = {
                "aceSku": rec.get("Ace SKU", ""),
                "retail": rec.get("Search Description", "")
                or rec.get("Item Description", ""),
                "status": rec.get("Item Status", ""),
            }
    chains = []
    ws = wb["Replacement Chains"]
    header = None
    for row in ws.iter_rows(values_only=True):
        vals = ["" if v is None else str(v).strip() for v in row]
        if header is None:
            header = vals
            continue
        rec = dict(zip(header, vals))
        if rec.get("Part-Number"):
            chains.append(
                {
                    "marketing": rec.get("Marketing Number", ""),
                    "part": rec.get("Part-Number", ""),
                    "desc": rec.get("Description", ""),
                    "aceSku": rec.get("ACE SKU #", ""),
                    "upc": rec.get("UPC", ""),
                }
            )
    return upc_map, chains


def main():
    price_csv = SOURCE / "Dealer_Price_File.csv"
    master_xlsx = SOURCE / "STIHL_SKU_Master_Listing.xlsx"
    for p in (price_csv, master_xlsx):
        if not p.exists():
            sys.exit(f"missing source file: {p}")

    upc_map, chains = load_master(master_xlsx)
    rows = list(csv.DictReader(open(price_csv, newline="", encoding="utf-8-sig")))

    models = {}
    bars = []
    accessories = []
    for r in rows:
        cat = r["Category"].strip()
        mat = r["STIHL Material Number"].strip()
        desc = r["Material Description"].strip()
        msrp = r["MSRP"].strip()
        upc = r["UPC"].strip()
        ace = r["ACE SKU"].strip()
        if not mat or not msrp:
            continue
        master = upc_map.get(upc, {})
        ace = ace or master.get("aceSku", "")

        if cat == "2BR":
            bm = BAR_RE.search(desc)
            bars.append(
                {
                    "part": dash_part(mat),
                    "desc": desc,
                    "lengthIn": int(bm.group(2)) if bm else None,
                    "msrp": float(msrp),
                    "aceSku": ace,
                }
            )
            continue

        if cat not in UNIT_CATEGORIES:
            continue

        model, ptype, bar_in, chain = parse_unit(desc)
        retail = master.get("retail", "")
        # dealer descriptions are 40-char truncated; recover saw bar length
        # from the master listing's retail description when missing
        if bar_in is None and cat in ("0CS", "0LB"):
            rm = re.search(r"\b(\d{2})\s?in\b", retail)
            if rm and 10 <= int(rm.group(1)) <= 36:
                bar_in = int(rm.group(1))
        nickname = ""
        for n in NICKNAMES:
            if n.lower() in (retail + " " + desc).lower():
                nickname = n.upper()
                break
        key = (cat, model)
        grp = models.setdefault(
            key,
            {
                "id": f"{cat}:{model}",
                "model": model,
                "nickname": nickname,
                "category": cat,
                "categoryName": UNIT_CATEGORIES[cat][0],
                "signCategory": UNIT_CATEGORIES[cat][1],
                "productType": ptype,
                "variants": [],
            },
        )
        if nickname and not grp["nickname"]:
            grp["nickname"] = nickname
        grp["variants"].append(
            {
                "material": mat,
                "materialDash": dash_part(mat),
                "desc": desc,
                "retail": retail,
                "barIn": bar_in,
                "chain": chain,
                "msrp": float(msrp),
                "upc": upc,
                "aceSku": ace,
                "status": master.get("status", ""),
            }
        )

    model_list = sorted(models.values(), key=lambda g: (g["category"], g["model"]))
    for g in model_list:
        g["variants"].sort(key=lambda v: (v["barIn"] or 0, v["msrp"]))

    data = {
        "categories": [
            {"code": c, "name": n[0], "signLabel": n[1]}
            for c, n in sorted(UNIT_CATEGORIES.items())
        ],
        "models": model_list,
        "chains": chains,
        "bars": sorted(bars, key=lambda b: (b["lengthIn"] or 0, b["part"])),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "// Generated by tools/build_data.py — do not edit by hand.\n"
        "window.SIGN_DATA = " + json.dumps(data, indent=1) + ";\n",
        encoding="utf-8",
    )
    n_var = sum(len(g["variants"]) for g in model_list)
    print(f"wrote {OUT}: {len(model_list)} models, {n_var} variants, "
          f"{len(chains)} chains, {len(bars)} bars")


if __name__ == "__main__":
    main()
